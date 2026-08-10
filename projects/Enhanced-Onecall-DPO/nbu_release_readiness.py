#!/usr/bin/env python3
"""
NetBackup Release Readiness Assessment
Fetches release notes for a target NBU version, reads existing site infra from
a Process_jobs Excel report, scores feature relevance, and emits JSON + CSV
suitable for Power BI data cards.

Usage:
    python nbu_release_readiness.py --excel <path_to_xlsx> [--version 11.2] [--output-dir .]
"""

import argparse
import csv
import json
import re
import sys
from datetime import datetime
from pathlib import Path

import openpyxl
import requests
from bs4 import BeautifulSoup

# ---------------------------------------------------------------------------
# Release note page definitions  (extend for future versions)
# ---------------------------------------------------------------------------
RELEASE_CATALOG = {
    "11.2": {
        "base_url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0",
        "features_url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172450863-171368441",
        "features": [
            {
                "id": "cohesity-terminology",
                "title": "Changes in Cohesity terminology",
                "category": "General",
                "description": "Veritas products now use Cohesity branding. UI labels, documentation, and command outputs reflect new names.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v149833249-171368441",
                "tags": ["all"],
                "impact": "Medium",
            },
            {
                "id": "restful-apis",
                "title": "RESTful APIs included in NetBackup 11.2",
                "category": "API / Automation",
                "description": "New and updated REST APIs for job management, policy configuration, and reporting.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172451280-171368441",
                "tags": ["all"],
                "impact": "High",
            },
            {
                "id": "stig-msdp",
                "title": "Support for STIG compliance for NetBackup MSDP",
                "category": "Security",
                "description": "MSDP storage servers can now be configured in STIG-compliant mode for government/regulated environments.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172450943-171368441",
                "tags": ["msdp", "worm", "security"],
                "impact": "High",
            },
            {
                "id": "archive-tier-verification",
                "title": "Support for verification of images stored in the archive tier",
                "category": "Storage",
                "description": "NetBackup can now verify the integrity of backup images offloaded to the archive tier.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172630778-171368441",
                "tags": ["msdp", "disk"],
                "impact": "Medium",
            },
            {
                "id": "msdp-dedup-performance",
                "title": "Enhanced MSDP deduplication and performance",
                "category": "Storage",
                "description": "Improved deduplication ratios and throughput for MSDP storage pools.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172630789-171368441",
                "tags": ["msdp", "worm", "disk"],
                "impact": "High",
            },
            {
                "id": "helios-integration",
                "title": "NetBackup integration with Helios",
                "category": "Management",
                "description": "Primary servers can be registered with Cohesity Helios for centralised cloud-based management.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172477643-171368441",
                "tags": ["all"],
                "impact": "High",
            },
            {
                "id": "worm-local-cache",
                "title": "Local WORM cache support for MSDP cloud",
                "category": "Storage",
                "description": "MSDP cloud storage can use a local WORM cache to improve restore performance for immutable backups.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172631120-171368441",
                "tags": ["msdp", "worm", "cloud"],
                "impact": "High",
            },
            {
                "id": "accelerator-oci",
                "title": "Accelerator support for Oracle Cloud Infrastructure (OCI)",
                "category": "Cloud",
                "description": "NetBackup Accelerator can now be used to back up data directly to OCI object storage.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172596506-171368441",
                "tags": ["cloud", "oracle"],
                "impact": "Medium",
            },
            {
                "id": "dynamic-multistream",
                "title": "Dynamic multi-streaming for Standard and Catalog policies",
                "category": "Backup",
                "description": "Standard and NBU-Catalog policies can now automatically scale data streams for faster backups.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172596511-171368441",
                "tags": ["standard", "nbu-catalog"],
                "impact": "High",
            },
            {
                "id": "cloud-autoresume",
                "title": "Auto-resume for Cloud object store backups",
                "category": "Cloud",
                "description": "Cloud object store jobs can automatically resume after transient failures without restarting from scratch.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172596516-171368441",
                "tags": ["cloud"],
                "impact": "Medium",
            },
            {
                "id": "cloud-scale-enhancements",
                "title": "Cloud Scale enhancements and deployment updates",
                "category": "Cloud",
                "description": "Updates to Cloud Scale Technology deployment for AWS and Azure including Terraform improvements.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172702917-171368441",
                "tags": ["cloud"],
                "impact": "Medium",
            },
            {
                "id": "parallel-restore-cloud-vm",
                "title": "Parallel restore for cloud virtual machines",
                "category": "Restore",
                "description": "Cloud-hosted VMs can now be restored in parallel, significantly reducing RTO.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172703229-171368441",
                "tags": ["vmware", "cloud"],
                "impact": "High",
            },
            {
                "id": "azure-deny-all",
                "title": "Support for protecting Azure disks with DENY_ALL network policy",
                "category": "Cloud",
                "description": "NetBackup can now back up Azure managed disks even when the DENY_ALL network policy is enforced.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172703368-171368441",
                "tags": ["cloud", "vmware"],
                "impact": "Medium",
            },
            {
                "id": "file-hash-threat-detection",
                "title": "File hash calculation for threat detection",
                "category": "Security",
                "description": "NetBackup can compute file hashes during backup for use with anomaly detection and threat intelligence.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172703373-171368441",
                "tags": ["all", "security"],
                "impact": "High",
            },
            {
                "id": "job-resiliency-oracle-sql",
                "title": "Job resiliency for Oracle and MS-SQL-Server policies",
                "category": "Database",
                "description": "Backup jobs for Oracle and MS-SQL-Server can now recover from transient errors without full restarts.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172704007-171368441",
                "tags": ["oracle", "ms-sql-server"],
                "impact": "High",
            },
            {
                "id": "kvm-enhancements",
                "title": "Enhancements in KVM support",
                "category": "Virtualization",
                "description": "Improved support for KVM virtual machines including new backup and restore capabilities.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172704058-171368441",
                "tags": ["kvm"],
                "impact": "Low",
            },
            {
                "id": "sap-hana-intelligent-policy",
                "title": "SAP HANA intelligent policy",
                "category": "Database",
                "description": "New intelligent policy type for automated SAP HANA backup management.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172704070-171368441",
                "tags": ["sap"],
                "impact": "Low",
            },
            {
                "id": "role-elevation",
                "title": "Role elevation feature",
                "category": "Security",
                "description": "Users can now temporarily elevate their RBAC role for specific administrative tasks.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172704185-171368441",
                "tags": ["all", "security"],
                "impact": "High",
            },
            {
                "id": "kms-webui",
                "title": "KMS configuration and management using web UI",
                "category": "Security",
                "description": "Key Management Service (KMS) can now be fully configured and managed through the NetBackup web UI.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172704501-171368441",
                "tags": ["all", "security", "msdp", "worm"],
                "impact": "Medium",
            },
            {
                "id": "troubleshooter-webui",
                "title": "Troubleshooter in web UI",
                "category": "Web UI",
                "description": "New guided troubleshooting wizard in the web UI to diagnose common job failure scenarios.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172704517-171368441",
                "tags": ["all"],
                "impact": "Medium",
            },
            {
                "id": "eeb-management-webui",
                "title": "Enhancements in EEB management web UI",
                "category": "Web UI",
                "description": "Improved Emergency Engineering Binary (EEB) listing, installation status, and management in the web UI.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172704518-171368441",
                "tags": ["all"],
                "impact": "Medium",
            },
            {
                "id": "exchange-recovery-webui",
                "title": "MS Exchange Recovery from NetBackup Web UI",
                "category": "Web UI",
                "description": "Microsoft Exchange mailbox and database restores can now be initiated directly from the web UI.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172627819-171368441",
                "tags": ["ms-exchange-server", "ms-windows"],
                "impact": "Medium",
            },
            {
                "id": "log-collection-webui",
                "title": "Enhanced Log Collection in NetBackup Web UI",
                "category": "Web UI",
                "description": "Log collection for support cases can now be triggered and downloaded through the web UI.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172477452-171368441",
                "tags": ["all"],
                "impact": "Medium",
            },
            {
                "id": "backup-selections-tab",
                "title": "Backup Selections Tab in Policies in NetBackup Web UI",
                "category": "Web UI",
                "description": "A dedicated Backup Selections tab is now available when creating or editing policies in the web UI.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172477465-171368441",
                "tags": ["all"],
                "impact": "Low",
            },
            {
                "id": "job-priority-activity-monitor",
                "title": "Change job priority for queued jobs in Activity Monitor",
                "category": "Web UI",
                "description": "Administrators can now adjust the priority of queued jobs in real time from the Activity Monitor.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172473980-171368441",
                "tags": ["all"],
                "impact": "Medium",
            },
            {
                "id": "enhanced-report-time",
                "title": "Enhanced report time selection and control",
                "category": "Web UI",
                "description": "Reporting views now support custom time ranges and improved filtering.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172477619-171368441",
                "tags": ["all"],
                "impact": "Low",
            },
            {
                "id": "mongodb-ops-manager",
                "title": "Enhanced Backup Capabilities for MongoDB Ops Manager",
                "category": "Database",
                "description": "New features for MongoDB Ops Manager backups including incremental support.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172476607-171368441",
                "tags": ["mongodb"],
                "impact": "Low",
            },
            {
                "id": "java-ui-eol",
                "title": "EOL of NetBackup Administration Console (Java UI) starting with 11.2",
                "category": "Admin - Breaking Change",
                "description": "The legacy Java-based Administration Console is end-of-life in 11.2. All administration must migrate to the web UI.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172705033-171368441",
                "tags": ["all"],
                "impact": "Critical",
            },
            {
                "id": "k8s-flexible-version",
                "title": "Flexible Version Compatibility for Kubernetes Operator",
                "category": "Kubernetes",
                "description": "The NetBackup Kubernetes operator now supports a broader range of Kubernetes API versions.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172627734-171368441",
                "tags": ["kubernetes"],
                "impact": "High",
            },
            {
                "id": "shutdown-commands-deprecated",
                "title": "Several shutdown commands to be deprecated in a future release",
                "category": "Admin - Breaking Change",
                "description": "bpdown and related shutdown commands are marked for deprecation. Migrate to supported alternatives.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v114644710-171368441",
                "tags": ["all"],
                "impact": "Medium",
            },
            {
                "id": "support-additions-11.2",
                "title": "NetBackup 11.2 support additions and changes",
                "category": "Support",
                "description": "New OS, hardware, and platform support added in 11.2 including updated compatibility matrices.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172451039-171368441",
                "tags": ["all"],
                "impact": "Medium",
            },
            {
                "id": "cloud-config-update",
                "title": "Update cloud configuration file after install/upgrade to NetBackup 11.2",
                "category": "Cloud - Action Required",
                "description": "After upgrading to 11.2, the cloud configuration file must be updated on the primary and media servers immediately.",
                "url": "https://docs.cohesity.com/docs/netbackup/11.2/103228346-171368441-0/v172451203-171368441",
                "tags": ["cloud", "all"],
                "impact": "High",
            },
        ],
    }
}

# ---------------------------------------------------------------------------
# Tag-to-policy/storage keyword mapping for relevance scoring
# ---------------------------------------------------------------------------
TAG_KEYWORDS = {
    "vmware":          ["vmware", "vm", "vsphere"],
    "kubernetes":      ["kubernetes", "k8s", "ocp", "openshift", "container"],
    "standard":        ["standard"],
    "ms-windows":      ["ms-windows", "windows"],
    "ms-sql-server":   ["ms-sql-server", "mssql", "sql server"],
    "ms-exchange-server": ["ms-exchange-server", "exchange"],
    "oracle":          ["oracle"],
    "ndmp":            ["ndmp", "nas", "universal-share"],
    "nbu-catalog":     ["nbu-catalog", "catalog"],
    "sybase":          ["sybase"],
    "sap":             ["sap"],
    "mongodb":         ["mongodb"],
    "kvm":             ["kvm"],
    "msdp":            ["msdp", "diskpool", "puredisk"],
    "worm":            ["worm"],
    "disk":            ["disk", "diskpool"],
    "cloud":           ["cloud", "aws", "azure", "gcp", "oci"],
    "security":        ["all"],
    "all":             ["all"],
}


def read_infra_from_excel(xlsx_path: str) -> dict:
    """Extract site infrastructure profile from a Process_jobs Excel report."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)

    infra = {
        "report_date": None,
        "company": None,
        "master_server": None,
        "current_version": None,
        "servers": [],
        "policy_types": {},
        "storage_types": set(),
        "has_worm": False,
        "has_msdp": False,
        "has_tape": False,
        "has_cloud": False,
        "has_kubernetes": False,
        "has_slp": False,
        "job_count": 0,
    }

    # Info Page
    if "Info Page" in wb.sheetnames:
        for row in wb["Info Page"].iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            val = str(row[0])
            if "Processing run on:" in val:
                infra["report_date"] = val.replace("Processing run on:", "").strip()
            elif "Company Name:" in val:
                infra["company"] = val.replace("Company Name:", "").strip()
            elif "Master Server Name:" in val:
                infra["master_server"] = val.replace("Master Server Name:", "").strip()
            elif "Jobs contained in this report:" in val:
                m = re.search(r"\d+", val)
                if m:
                    infra["job_count"] = int(m.group())

    # Server_Versions
    if "Server_Versions" in wb.sheetnames:
        for row in wb["Server_Versions"].iter_rows(min_row=2, values_only=True):
            if row[0] and row[1] and row[2]:
                infra["servers"].append({"role": row[0], "version": row[1], "hostname": row[2]})
                if infra["current_version"] is None and row[0] == "MASTER":
                    infra["current_version"] = row[1]

    # Clients – policy types
    if "Clients" in wb.sheetnames:
        from collections import Counter
        pt_counter = Counter()
        for row in wb["Clients"].iter_rows(min_row=2, values_only=True):
            if row and row[2]:
                pt_counter[str(row[2])] += 1
        infra["policy_types"] = dict(pt_counter)

    # Storage_Units
    if "Storage_Units" in wb.sheetnames:
        for row in wb["Storage_Units"].iter_rows(min_row=2, values_only=True):
            if not row or row[0] is None:
                continue
            label = str(row[0]).lower()
            stype = str(row[1]).lower() if row[1] else ""
            sub   = str(row[2]).lower() if row[2] else ""
            worm_capable = str(row[15]).lower() if len(row) > 15 and row[15] else ""
            infra["storage_types"].add(stype)
            if "media manager" in stype:
                infra["has_tape"] = True
            if "disk" in stype and ("diskpool" in sub or "puredisk" in label or "msdp" in label):
                infra["has_msdp"] = True
            if "worm" in label or "worm" in worm_capable or worm_capable == "yes":
                infra["has_worm"] = True

    # Detect Kubernetes and SLP
    infra["has_kubernetes"] = "Kubernetes" in infra["policy_types"]
    if "SLP_Report_Summary" in wb.sheetnames:
        for row in wb["SLP_Report_Summary"].iter_rows(min_row=2, values_only=True):
            if row and any(c for c in row):
                infra["has_slp"] = True
                break

    infra["storage_types"] = sorted(infra["storage_types"])
    return infra


def score_feature(feature: dict, infra: dict) -> tuple[str, str]:
    """
    Return (relevance_label, reason) for a feature given the site infra.
    Relevance: Critical | High | Medium | Low | Not Applicable
    """
    tags = [t.lower() for t in feature.get("tags", [])]
    policy_types_lower = {pt.lower() for pt in infra["policy_types"]}

    matched_reasons = []

    for tag in tags:
        if tag == "all":
            matched_reasons.append("Applies to all NetBackup environments")
            break
        if tag == "vmware" and "vmware" in policy_types_lower:
            matched_reasons.append(f"Site has VMware policies ({infra['policy_types'].get('VMware', 0)} clients)")
        if tag == "kubernetes" and infra["has_kubernetes"]:
            matched_reasons.append(f"Site has Kubernetes policies ({infra['policy_types'].get('Kubernetes', 0)} clients)")
        if tag == "standard" and "standard" in policy_types_lower:
            matched_reasons.append(f"Site has Standard policies ({infra['policy_types'].get('Standard', 0)} policies)")
        if tag == "nbu-catalog" and "nbu-catalog" in policy_types_lower:
            matched_reasons.append("Site has NBU-Catalog policy")
        if tag == "ndmp" and ("ndmp" in policy_types_lower or "universal-share" in policy_types_lower):
            matched_reasons.append("Site has NDMP/Universal-Share policies")
        if tag == "ms-windows" and "ms-windows" in policy_types_lower:
            matched_reasons.append("Site has MS-Windows policies")
        if tag == "ms-sql-server" and "ms-sql-server" in policy_types_lower:
            matched_reasons.append("Site has MS-SQL-Server policies")
        if tag == "oracle" and "oracle" in policy_types_lower:
            matched_reasons.append("Site has Oracle policies")
        if tag == "sybase" and "sybase" in policy_types_lower:
            matched_reasons.append("Site has Sybase policies")
        if tag in ("msdp", "disk") and infra["has_msdp"]:
            matched_reasons.append("Site uses MSDP disk pool storage")
        if tag == "worm" and infra["has_worm"]:
            matched_reasons.append("Site has WORM-capable storage")
        if tag == "cloud" and infra["has_cloud"]:
            matched_reasons.append("Site has cloud storage units")

    if not matched_reasons:
        return "Not Applicable", "No matching workloads or storage detected at this site"

    # Inherit the feature's own impact rating, but cap at the feature's stated impact
    relevance = feature.get("impact", "Medium")
    return relevance, "; ".join(dict.fromkeys(matched_reasons))  # dedupe while keeping order


def build_output_rows(version: str, infra: dict) -> list[dict]:
    catalog = RELEASE_CATALOG.get(version)
    if not catalog:
        raise ValueError(f"No release catalog entry for version {version}")

    rows = []
    run_ts = datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ")

    for feat in catalog["features"]:
        relevance, reason = score_feature(feat, infra)
        rows.append(
            {
                "run_timestamp":    run_ts,
                "site_company":     infra["company"],
                "master_server":    infra["master_server"],
                "current_version":  infra["current_version"],
                "target_version":   version,
                "feature_id":       feat["id"],
                "feature_title":    feat["title"],
                "category":         feat["category"],
                "impact":           feat["impact"],
                "site_relevance":   relevance,
                "relevance_reason": reason,
                "description":      feat["description"],
                "doc_url":          feat["url"],
                "policy_types_on_site": "; ".join(
                    f"{k}:{v}" for k, v in sorted(infra["policy_types"].items())
                ),
                "has_msdp":   infra["has_msdp"],
                "has_worm":   infra["has_worm"],
                "has_tape":   infra["has_tape"],
                "has_k8s":    infra["has_kubernetes"],
                "has_slp":    infra["has_slp"],
                "job_count":  infra["job_count"],
                "server_count": len(infra["servers"]),
            }
        )
    return rows


def write_csv(rows: list[dict], path: Path) -> None:
    if not rows:
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        writer.writeheader()
        writer.writerows(rows)


def write_json(rows: list[dict], infra: dict, version: str, path: Path) -> None:
    payload = {
        "metadata": {
            "generated_at": datetime.utcnow().strftime("%Y-%m-%dT%H:%M:%SZ"),
            "target_version": version,
            "release_notes_url": RELEASE_CATALOG[version]["features_url"],
            "site_company": infra["company"],
            "master_server": infra["master_server"],
            "current_version": infra["current_version"],
            "server_count": len(infra["servers"]),
            "job_count": infra["job_count"],
            "policy_types": infra["policy_types"],
            "storage_profile": {
                "has_msdp": infra["has_msdp"],
                "has_worm": infra["has_worm"],
                "has_tape": infra["has_tape"],
                "has_cloud": infra["has_cloud"],
                "has_kubernetes": infra["has_kubernetes"],
                "has_slp": infra["has_slp"],
            },
        },
        "features": rows,
    }
    with open(path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, default=str)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(description="NetBackup Release Readiness Assessment")
    parser.add_argument(
        "--excel",
        required=True,
        help="Path to the Process_jobs Excel report (.xlsx)",
    )
    parser.add_argument(
        "--version",
        default="11.2",
        help="Target NetBackup version (default: 11.2)",
    )
    parser.add_argument(
        "--output-dir",
        default=".",
        help="Directory for output files (default: current dir)",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.excel).resolve()
    if not xlsx_path.exists():
        sys.exit(f"ERROR: Excel file not found: {xlsx_path}")

    if args.version not in RELEASE_CATALOG:
        sys.exit(
            f"ERROR: Version '{args.version}' not in catalog. "
            f"Available: {list(RELEASE_CATALOG.keys())}"
        )

    output_dir = Path(args.output_dir).resolve()
    output_dir.mkdir(parents=True, exist_ok=True)

    print(f"[1/3] Reading infra from: {xlsx_path.name}")
    infra = read_infra_from_excel(str(xlsx_path))
    print(
        f"      Site : {infra['company']} | Master: {infra['master_server']} | "
        f"Version: {infra['current_version']} | Jobs: {infra['job_count']}"
    )
    print(f"      Policy types: {infra['policy_types']}")

    print(f"[2/3] Evaluating {len(RELEASE_CATALOG[args.version]['features'])} "
          f"NetBackup {args.version} features against site profile...")
    rows = build_output_rows(args.version, infra)

    # Summary
    from collections import Counter
    rel_counts = Counter(r["site_relevance"] for r in rows)
    print(f"      Relevance breakdown: {dict(rel_counts)}")

    safe_master = re.sub(r"[^\w.-]", "_", infra["master_server"] or "unknown")
    ts = datetime.utcnow().strftime("%Y%m%d")
    stem = f"nbu{args.version.replace('.','')}_readiness_{safe_master}_{ts}"

    csv_path  = output_dir / f"{stem}.csv"
    json_path = output_dir / f"{stem}.json"

    print(f"[3/3] Writing outputs to {output_dir}")
    write_csv(rows, csv_path)
    write_json(rows, infra, args.version, json_path)
    print(f"      CSV  → {csv_path}")
    print(f"      JSON → {json_path}")
    print("Done.")


if __name__ == "__main__":
    main()
